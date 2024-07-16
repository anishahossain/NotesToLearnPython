import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_file(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # range(1, 4) will only iterate over rows 1 2 3 hence we add plus one so all rows are iterated
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_cell = sheet.cell(row, 4) #this will create a new column in our excel file
        corrected_cell.value = corrected_price

    corrected_cells_title = sheet['d1']
    corrected_cells_title.value = 'corrected price'

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    # we pass values as such to specify the input values for the chart (line21)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2') #e2 (or any coordinate we place here that we want) is the position where the chart will be added in excel file

    wb.save(filename) # this creates a new excel file with updates we made so that original file is not over written
    # / saves the updates to the filename mentioned
